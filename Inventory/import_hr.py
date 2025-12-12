import tkinter as tk
from tkinter import messagebox, filedialog
from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

import shared_functions as sf

STYLE_TEMPLATE_ROW = 2  # first data row under headers

# Columns to hide before saving
HIDDEN_COLUMNS = ["B", "D", "E", "F", "G", "K", "L", "N", "P", "Q", "R", "S", "T", "U", "V", "W"]


def copy_cell_styles(ws, src_row: int, dst_row: int, cols: list[str]) -> None:
    """
    Copy the entire style object from src_row to dst_row for the given columns.
    Using dst._style = copy(src._style) avoids StyleProxy / hashability issues.
    """
    for col in cols:
        src = ws[f"{col}{src_row}"]
        dst = ws[f"{col}{dst_row}"]
        dst._style = copy(src._style)


def sort_worksheet_by_column_a(ws) -> None:
    """
    Sort the data rows (row 2 .. max_row) by column A (ascending).
    """
    max_row = ws.max_row
    max_col = ws.max_column

    data_rows = list(
        ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True)
    )

    def key_func(row):
        val = row[0]  # column A
        try:
            return int(val)
        except (TypeError, ValueError):
            return float("inf")

    data_rows.sort(key=key_func)

    for r_idx, row_values in enumerate(data_rows, start=2):
        for c_idx, value in enumerate(row_values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def normalize_serial(val) -> str:
    """
    Simple normalization for serial numbers: string, stripped, uppercased.
    """
    if val is None:
        return ""
    return str(val).strip().upper()


def apply_hidden_column_styles_and_hide(ws) -> None:
    """
    For each hidden column we care about:
      - Copy the style (font/size/number_format/etc.) from STYLE_TEMPLATE_ROW
        to all data rows.
      - Mark the column as hidden.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for col_letter in HIDDEN_COLUMNS:
        try:
            col_idx = column_index_from_string(col_letter)
        except ValueError:
            continue

        if col_idx > max_col:
            continue

        src = ws[f"{col_letter}{STYLE_TEMPLATE_ROW}"]
        for row in range(2, max_row + 1):
            dst = ws[f"{col_letter}{row}"]
            dst._style = copy(src._style)

        ws.column_dimensions[col_letter].hidden = True


def apply_hr_row_to_existing(
    ws_existing,
    ws_new,
    src_row: int,
    dst_row: int,
    *,
    existing_max_col: int,
    new_max_col: int,
) -> None:
    """
    Copy HR values from ws_new[src_row] into ws_existing[dst_row], shifting by +1 column,
    but do NOT overwrite:
      - Column I (Loc)
      - Columns X-AA (In/Out, LOE, Last Scanned, Last Verified)

    Column A is managed separately and is not written here.

    Mapping: new col c -> existing col (c + 1)
    """
    max_copy_cols = min(new_max_col, existing_max_col - 1)  # dest is c+1

    # Destination columns to skip by numeric index
    # I = 9
    # X..AA = 24..27
    skip_dest_cols = {9, 24, 25, 26, 27}

    for c in range(1, max_copy_cols + 1):
        dest_col = c + 1
        if dest_col in skip_dest_cols:
            continue
        val = ws_new.cell(row=src_row, column=c).value
        ws_existing.cell(row=dst_row, column=dest_col, value=val)


def run_import_inventory(root: tk.Misc | None = None):
    """
    Import a new inventory sheet into the existing inventory.

    Matching rule:
      - New sheet: Asset Id in column B, Serial in column N.
      - Existing sheet: Asset Id in column C, Serial in column O.

    For each new-row (in order, 1-based index i):
      1) Try match by (AssetId, Serial).
         - If match:
             * Set existing row's column A = i.
             * Overwrite existing row with HR values for all mapped columns
               EXCEPT: A, I, and X-AA.
      2) If no match by Asset+Serial:
         - Look for a 50xxx row where Serial matches (by normalized serial only).
             * If found:
                 - Use that row:
                     A = i
                     Overwrite existing row with HR values for all mapped columns
                     EXCEPT: A, I, and X-AA.
             * Else:
                 - Append new row at bottom:
                     A = i
                     Copy HR values shifted one column to the right,
                     EXCEPT: I and X-AA (these are left blank).
                     X-AA are explicitly blanked for brand new items.

    After processing all new rows:
      - Any original existing rows with A < 50000 and not matched:
          * Set A to NEXT AVAILABLE 50xxx number (re-using gaps).
          * -> "removed by HR".

      - Sort by column A.
      - Apply hidden-column styling + hide them.
      - Check for duplicate Asset Ids (FULL string; ignore N/A and Other Equipment).
      - Save and show summary with counts, lists, and duplicate report.
    """
    owns_root = False
    if root is None:
        root = tk.Tk()
        root.withdraw()
        owns_root = True

    # Select existing inventory file
    existing_path_str = sf.select_inventory_excel_file(
        root=root, title="Select EXISTING Inventory Excel File"
    )
    if not existing_path_str:
        if owns_root:
            root.destroy()
        return
    existing_path = Path(existing_path_str)

    # Select new inventory file (HR export)
    new_path_str = filedialog.askopenfilename(
        parent=root,
        title="Select NEW Inventory Excel File (HR Export)",
        filetypes=[("Excel files", "*.xlsx *.xls")],
    )
    if not new_path_str:
        if owns_root:
            root.destroy()
        return
    new_path = Path(new_path_str)

    # Load existing workbook
    try:
        wb_existing = load_workbook(existing_path)
    except Exception as e:
        messagebox.showerror("Load error", f"Failed to open existing workbook:\n{e}", parent=root)
        if owns_root:
            root.destroy()
        return

    ws_existing = wb_existing[wb_existing.sheetnames[0]]
    existing_max_col = ws_existing.max_column

    # Load new workbook
    try:
        wb_new = load_workbook(new_path, data_only=True)
    except Exception as e:
        messagebox.showerror("Load error", f"Failed to open new (import) workbook:\n{e}", parent=root)
        if owns_root:
            root.destroy()
        return

    ws_new = wb_new[wb_new.sheetnames[0]]

    # Build lookup for existing rows by (normalized_asset, normalized_serial)
    existing_map: dict[tuple[str, str], int] = {}
    original_max_row = ws_existing.max_row
    original_A_values: dict[int, int] = {}

    # Lookup from normalized serial -> 50xxx row (serial-only promotion)
    serial_to_50xxx_row: dict[str, int] = {}

    for r in range(2, original_max_row + 1):
        a_val_raw = ws_existing[f"A{r}"].value
        try:
            a_val = int(str(a_val_raw))
        except (TypeError, ValueError):
            a_val = None
        if a_val is not None:
            original_A_values[r] = a_val

        asset_val = ws_existing.cell(row=r, column=3).value  # C
        serial_val = ws_existing.cell(row=r, column=15).value  # O

        asset_norm = sf.normalize(asset_val) if asset_val is not None else ""
        serial_norm = normalize_serial(serial_val)

        if asset_norm or serial_norm:
            existing_map[(asset_norm, serial_norm)] = r

        if a_val is not None and 50000 <= a_val < 51000 and serial_norm:
            serial_to_50xxx_row[serial_norm] = r

    matched_existing_rows: set[int] = set()

    # Collect used 50xxx numbers so we can reuse gaps
    used_50xxx: set[int] = set()
    for r, a_val in original_A_values.items():
        if 50000 <= a_val < 51000:
            used_50xxx.add(a_val)

    next_50xxx = 50000

    def get_next_50xxx() -> int:
        nonlocal next_50xxx
        while next_50xxx in used_50xxx:
            next_50xxx += 1
        used_50xxx.add(next_50xxx)
        value = next_50xxx
        next_50xxx += 1
        return value

    # Counts and lists
    matched_regular_count = 0
    added_from_50xxx_count = 0
    brand_new_count = 0

    brand_new_items: list[str] = []
    added_from_50xxx_items: list[str] = []

    import_index = 0
    new_max_row = ws_new.max_row
    new_max_col = ws_new.max_column

    for r in range(2, new_max_row + 1):
        asset_val = ws_new.cell(row=r, column=2).value  # B
        serial_val = ws_new.cell(row=r, column=14).value  # N

        asset_norm = sf.normalize(asset_val) if asset_val is not None else ""
        serial_norm = normalize_serial(serial_val)

        if not asset_norm and not serial_norm:
            continue

        import_index += 1
        key = (asset_norm, serial_norm)

        if key in existing_map:
            # Exact match
            existing_row = existing_map[key]
            old_a = original_A_values.get(existing_row)

            ws_existing[f"A{existing_row}"].value = import_index
            matched_existing_rows.add(existing_row)

            # Overwrite with HR values (except A, I, X-AA)
            apply_hr_row_to_existing(
                ws_existing,
                ws_new,
                src_row=r,
                dst_row=existing_row,
                existing_max_col=existing_max_col,
                new_max_col=new_max_col,
            )

            if old_a is not None and 50000 <= old_a < 51000:
                added_from_50xxx_count += 1
                added_from_50xxx_items.append(
                    f"Row {existing_row}: Asset {asset_val!r} / Serial {serial_val!r} "
                    f"(was {old_a}, now {import_index})"
                )
            else:
                matched_regular_count += 1

        else:
            # No exact match: try serial-only promotion from 50xxx
            serial_row = serial_to_50xxx_row.get(serial_norm)

            if serial_row is not None:
                old_a = original_A_values.get(serial_row)

                ws_existing[f"A{serial_row}"].value = import_index
                matched_existing_rows.add(serial_row)

                # Overwrite with HR values (except A, I, X-AA)
                apply_hr_row_to_existing(
                    ws_existing,
                    ws_new,
                    src_row=r,
                    dst_row=serial_row,
                    existing_max_col=existing_max_col,
                    new_max_col=new_max_col,
                )

                added_from_50xxx_count += 1
                added_from_50xxx_items.append(
                    f"Row {serial_row}: Serial-only match, updated from HR. "
                    f"Asset {asset_val!r} / Serial {serial_val!r} (was {old_a}, now {import_index})"
                )

            else:
                # Brand new: append row
                dest_row = ws_existing.max_row + 1
                brand_new_count += 1

                # Clear any existing values in this row
                for c in range(1, existing_max_col + 1):
                    ws_existing.cell(row=dest_row, column=c, value=None)

                # Copy ENTIRE ROW styling from template
                cols_to_style = [get_column_letter(c) for c in range(1, existing_max_col + 1)]
                copy_cell_styles(ws_existing, STYLE_TEMPLATE_ROW, dest_row, cols_to_style)

                ws_existing[f"A{dest_row}"].value = import_index

                # Apply HR values (except I and X-AA)
                apply_hr_row_to_existing(
                    ws_existing,
                    ws_new,
                    src_row=r,
                    dst_row=dest_row,
                    existing_max_col=existing_max_col,
                    new_max_col=new_max_col,
                )

                # Ensure status / scan columns are blank for brand-new items
                for col_letter in ("X", "Y", "Z", "AA"):
                    ws_existing[f"{col_letter}{dest_row}"].value = ""

                brand_new_items.append(
                    f"Row {dest_row}: Asset {asset_val!r} / Serial {serial_val!r}"
                )

    # Mark removed-by-HR: original rows with A < 50000 and not matched
    removed_count = 0
    removed_items: list[str] = []

    for r in range(2, original_max_row + 1):
        if r in matched_existing_rows:
            continue

        a_raw = ws_existing[f"A{r}"].value
        try:
            a_val = int(str(a_raw))
        except (TypeError, ValueError):
            continue

        if a_val < 50000:
            new_50 = get_next_50xxx()
            ws_existing[f"A{r}"].value = new_50
            removed_count += 1

            asset_val = ws_existing.cell(row=r, column=3).value  # C
            serial_val = ws_existing.cell(row=r, column=15).value  # O
            removed_items.append(
                f"Row {r}: Asset {asset_val!r} / Serial {serial_val!r} -> {new_50}"
            )

    # Sort rows by column A
    sort_worksheet_by_column_a(ws_existing)

    # Apply hidden-column styling + hide them
    apply_hidden_column_styles_and_hide(ws_existing)

    # Duplicate Asset Id check (FULL asset ID, not normalized), ignore N/A and Other Equipment
    dup_groups: list[str] = []
    asset_map: dict[str, list[int]] = {}

    max_row_after = ws_existing.max_row
    for r in range(2, max_row_after + 1):
        raw_asset = ws_existing.cell(row=r, column=3).value  # C
        if raw_asset is None:
            continue
        raw_str = str(raw_asset).strip()
        if raw_str == "" or raw_str.upper() == "N/A" or raw_str.lower() == "other equipment":
            continue
        asset_map.setdefault(raw_str, []).append(r)

    for raw_asset, rows in asset_map.items():
        if len(rows) > 1:
            block = (
                f"Asset ID '{raw_asset}' appears in {len(rows)} rows:\n"
                + "\n".join(f"    row {rr}" for rr in rows)
            )
            dup_groups.append(block)

    # Save
    try:
        wb_existing.save(existing_path)
    except Exception as e:
        messagebox.showerror("Save error", f"Failed to save changes:\n{e}", parent=root)
        if owns_root:
            root.destroy()
        return

    # Summary
    summary_lines = [
        "Successfully imported new inventory.",
        "",
        f"No-change matches (existing non-50xxx items): {matched_regular_count}",
        f"Added items (matched existing 50xxx rows, including serial-only matches): {added_from_50xxx_count}",
        f"Brand new items (only in HR sheet): {brand_new_count}",
        f"Removed by HR (re-numbered to 50xxx): {removed_count}",
    ]

    if added_from_50xxx_items:
        summary_lines.append("")
        summary_lines.append("Added items (from 50xxx):")
        summary_lines.extend(f"  • {s}" for s in added_from_50xxx_items)

    if brand_new_items:
        summary_lines.append("")
        summary_lines.append("Brand new items:")
        summary_lines.extend(f"  • {s}" for s in brand_new_items)

    if removed_items:
        summary_lines.append("")
        summary_lines.append("Removed items (moved to 50xxx):")
        summary_lines.extend(f"  • {s}" for s in removed_items)

    if dup_groups:
        summary_lines.append("")
        summary_lines.append("⚠ Duplicate Asset IDs detected (by full Asset Id):")
        for block in dup_groups:
            summary_lines.append("  • " + block.replace("\n", "\n    "))
    else:
        summary_lines.append("")
        summary_lines.append("No duplicate Asset IDs detected (by full Asset Id).")

    messagebox.showinfo("Import complete", "\n".join(summary_lines), parent=root)

    if owns_root:
        root.destroy()


if __name__ == "__main__":
    run_import_inventory()